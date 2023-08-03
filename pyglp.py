from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from barcode import EAN13
from barcode.writer import ImageWriter
from PIL import Image as pil_img
import db_conection
import win32com.client as win32
import os
from datetime import datetime
import time




class PyGLP():
    user = ""
    path_logo = "./static/files/placas_de_preco/logo.jpg"
    celulas = {
        "bloco-1": {
            "titulo": "A2",
            "preço": "A4",
            "codigo": "A8",
            "logo": "D8"
        },
        "bloco-2": {
            "titulo": "F2",
            "preço": "F4",
            "codigo": "F8",
            "logo": "I8"
        },
        "bloco-3": {
            "titulo": "A12",
            "preço": "A14",
            "codigo": "A18",
            "logo": "D18"
        },
        "bloco-4": {
            "titulo": "F12",
            "preço": "F14",
            "codigo": "F18",
            "logo": "I18"
        },
        "bloco-5": {
            "titulo": "A22",
            "preço": "A24",
            "codigo": "A28",
            "logo": "D28"
        },
        "bloco-6": {
            "titulo": "F22",
            "preço": "F24",
            "codigo": "F28",
            "logo": "I28"
        },
        "bloco-7": {
            "titulo": "A32",
            "preço": "A34",
            "codigo": "A38",
            "logo": "D38"
        },
        "bloco-8": {
            "titulo": "F32",
            "preço": "F34",
            "codigo": "F38",
            "logo": "I38"
        },
        "bloco-9": {
            "titulo": "A42",
            "preço": "A44",
            "codigo": "A48",
            "logo": "D48"
        },
        "bloco-10": {
            "titulo": "F42",
            "preço": "F44",
            "codigo": "F48",
            "logo": "I48"
        }
    }
    dados = {
        "bloco-1": {
            "titulo": "",
            "preço": "",
            "codigo": "",
            "logo": ""
        },
        "bloco-2": {
            "titulo": "",
            "preço": "",
            "codigo": "",
            "logo": ""
        },
        "bloco-3": {
            "titulo": "",
            "preço": "",
            "codigo": "",
            "logo": ""
        },
        "bloco-4": {
            "titulo": "",
            "preço": "",
            "codigo": "",
            "logo": ""
        },
        "bloco-5": {
            "titulo": "",
            "preço": "",
            "codigo": "",
            "logo": ""
        },
        "bloco-6": {
            "titulo": "",
            "preço": "",
            "codigo": "",
            "logo": ""
        },
        "bloco-7": {
            "titulo": "",
            "preço": "",
            "codigo": "",
            "logo": ""
        },
        "bloco-8": {
            "titulo": "",
            "preço": "",
            "codigo": "",
            "logo": ""
        },
        "bloco-9": {
            "titulo": "",
            "preço": "",
            "codigo": "",
            "logo": ""
        },
        "bloco-10": {
            "titulo": "",
            "preço": "",
            "codigo": "",
            "logo": ""
        }
    }
    dict_values = None
    path_pdf_status = None
    long_path_pdf_saved = None
    short_path_pdf_saved = None
    pdf_file_name = None

    def __init__(self):
        pass

    def initialize(self, user, dict_values):
        self.user = user
        self.dict_values = dict_values
        try:
            os.makedirs('./static/files/placas_de_preco/{}'.format(self.user))
            os.makedirs('./static/files/placas_de_preco/{}/codes'.format(self.user))
        except:
            pass
        self.__clearValues()
        self.__setValues(self.dict_values)
        self.__startGenerate()

    def __clearValues(self):
        cont = 1
        while cont <= 10:
            self.dados[f"bloco-{cont}"]['titulo'] = ""
            self.dados[f"bloco-{cont}"]['preço'] = ""
            self.dados[f"bloco-{cont}"]['codigo'] = ""
            self.dados[f"bloco-{cont}"]['logo'] = ""
            cont = cont + 1

    def __setValues(self, list):
        #print(list)
        cont = 1
        for i in list:
            self.dados[f"bloco-{cont}"]['titulo'] = i['titulo']
            self.dados[f"bloco-{cont}"]['preço'] = i['preço']
            self.dados[f"bloco-{cont}"]['codigo'] = i['codigo']
            self.dados[f"bloco-{cont}"]['logo'] = f"{self.path_logo}"
            cont = cont + 1

    def __startGenerate(self):
        print("Start -  Generate Labels")
        wb = load_workbook(filename='./static/files/placas_de_preco/layout_glp.xlsx')
        ws = wb['main']
        i = 1
        while i <= 10:
            if self.dados[f"bloco-{i}"]['codigo'] != "":            
                #Title Product
                ws[self.celulas[f"bloco-{i}"]['titulo']] = self.dados[f"bloco-{i}"]['titulo']
                #Price Product
                ws[self.celulas[f"bloco-{i}"]['preço']] = self.dados[f"bloco-{i}"]['preço']
                #Code Product
                self.__generateBarCode(code=self.dados[f"bloco-{i}"]['codigo'], namefile=f"code-bloco-{i}")
                code_img = Image(f"./static/files/placas_de_preco/{self.user}/codes/code-bloco-{i}.png")
                anchor_code = self.celulas[f"bloco-{i}"]['codigo']
                ws.add_image(code_img, anchor_code) #dados[f"bloco-{i}"]['codigo']
                #Logo company
                logo = Image(f"{self.path_logo}")
                anchor_img = self.celulas[f"bloco-{i}"]['logo']
                ws.add_image(logo, anchor_img) #dados[f"bloco-{i}"]['preço']

                
                #Increment NOT REMOVE
            i = i + 1
        wb.save('./static/files/placas_de_preco/{}/new_page.xlsx'.format(self.user))
        print("End - Generate Labels")
        wb.close()
        del ws
        del wb
        self.__convert_xlsx_to_pdf()
    
    def __generateBarCode(self, code, namefile):
        if code != "":
            path_file = f"./static/files/placas_de_preco/{self.user}/codes/{namefile}"
            my_code = EAN13(code, writer=ImageWriter())
            my_code.save(path_file)


            img = pil_img.open(path_file+".png")
            new_height = 60
            new_width = int(new_height / img.height * img.width)
            new_size = img.resize((new_width, new_height))
            new_size.save(path_file+".png")
    
    def __convert_xlsx_to_pdf(self):
        atual_path = os.getcwdb().decode('utf-8')
        file_name = '{user}_{abreviate}.pdf'.format(user=self.user, abreviate=datetime.today().strftime('%d-%m-%Y_%H-%M-%S'))
        # /static/files/placas_de_preco/
        file_path = atual_path+'.\\static\\files\\placas_de_preco\\{user}\\new_page.xlsx'.format(user=self.user)
        save_path = atual_path+'.\\static\\files\\placas_de_preco\\{user}\\{file}'.format(user=self.user, file=file_name)

        excel = win32.Dispatch("Excel.Application")
        excel.Visible = False  # Para que o Excel não seja exibido durante a conversão
        wb = excel.Workbooks.Open(file_path)

        try:
            wb.ActiveSheet.ExportAsFixedFormat(0, save_path) #save_path  # 0 é o valor para PDF
            print(f"PDF Gerado com sucesso!")
            self.path_pdf_status = 'saved'
            self.pdf_file_name = file_name
        except Exception as e:
            print(f"Erro ao converter o arquivo: {e}")
        finally:
            wb.Close(SaveChanges=False)  # Fecha o arquivo sem salvar alterações
            #wb.Close(False)
            excel.Quit()
            excel = None
            del excel
            os.remove('./static/files/placas_de_preco/{}/new_page.xlsx'.format(self.user))

class DictGen():
    conexao = None
    def __init__(self):
        self.conexao = db_conection.ServerDB()

    def generateDictManual(self):
        new_code = ""
        cont = 1
        lista = []
        lista_dict = []
        while new_code.upper() != "N" or cont <= 10:
            new_code = input("{} - Digite um codigo (ou digite N para finalizar): ".format(cont))
            if new_code.upper() != "N":
                lista.append(new_code.replace("\n", ""))
            else:
                break
            cont = cont+1
        if len(lista) != 0:
            for i in lista:
                result = self.conexao.consulta_dict(i)
                if result != None:
                    lista_dict.append(result)
                else:
                    pass
            return lista_dict
        else:
            return print("Nenhum codigo foi detectado!")
        
    def generateDictAuto(self, lista_codes):
        lista = lista_codes
        lista_dict = []
        if len(lista) != 0:
            for i in lista:
                result = self.conexao.consulta_dict(i)
                if result != None:
                    lista_dict.append(result)
                else:
                    pass
            return lista_dict
        else:
            return print("Nenhum codigo foi detectado!")
